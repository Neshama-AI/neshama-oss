# -*- coding: utf-8 -*-
"""
文件异步任务
处理文件处理、向量化等耗时操作
"""

import logging
from celery import shared_task
from django.utils import timezone

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def vectorize_file_task(self, file_id: int):
    """
    文件向量化任务
    
    Args:
        file_id: 文件ID
    """
    from .models import FileStorage
    from .services import KnowledgeBaseService
    
    try:
        file_record = FileStorage.objects.get(id=file_id)
        
        # 检查文件是否有效
        if file_record.is_deleted:
            logger.info(f"文件已删除，跳过向量化: {file_id}")
            return {'status': 'skipped', 'reason': 'file_deleted'}
        
        if not file_record.agent:
            logger.info(f"文件无关联Agent，跳过向量化: {file_id}")
            return {'status': 'skipped', 'reason': 'no_agent'}
        
        # 更新状态为处理中
        file_record.vector_status = 'processing'
        file_record.save(update_fields=['vector_status'])
        
        # 获取文件内容
        from .services import get_storage_backend
        storage = get_storage_backend(file_record.storage_type)
        
        if file_record.storage_type == 'oss':
            content = storage.download(file_record.oss_key)
        else:
            content = storage.download(file_record.storage_path)
        
        # 根据文件类型提取文本
        text_content = extract_text_from_file(content, file_record.file_extension)
        
        if not text_content:
            raise ValueError("无法提取文本内容")
        
        # 调用向量化服务
        vector_result = call_vector_service(
            text=text_content,
            metadata={
                'file_id': file_record.id,
                'file_name': file_record.original_name,
                'agent_id': file_record.agent.id if file_record.agent else None,
            }
        )
        
        # 更新向量化结果
        KnowledgeBaseService.update_vectorization_result(
            file_record=file_record,
            vector_id=vector_result.get('vector_id', ''),
            status='completed'
        )
        
        logger.info(f"文件向量化完成: {file_id}, vector_id: {vector_result.get('vector_id')}")
        return {
            'status': 'success',
            'file_id': file_id,
            'vector_id': vector_result.get('vector_id')
        }
        
    except FileStorage.DoesNotExist:
        logger.error(f"文件不存在: {file_id}")
        return {'status': 'error', 'reason': 'file_not_found'}
        
    except Exception as exc:
        logger.error(f"文件向量化失败: {file_id}, error: {exc}")
        
        # 更新错误状态
        try:
            file_record = FileStorage.objects.get(id=file_id)
            KnowledgeBaseService.update_vectorization_result(
                file_record=file_record,
                status='failed',
                error=str(exc)
            )
        except Exception:
            pass
        
        # 重试
        raise self.retry(exc=exc)


@shared_task
def cleanup_expired_uploads():
    """清理过期的分片上传"""
    from .models import ChunkedUpload
    from .services import ChunkedUploadService
    
    expired_uploads = ChunkedUpload.objects.filter(
        expires_at__lt=timezone.now(),
        status__in=['pending', 'uploading', 'failed']
    )
    
    cleaned = 0
    for upload in expired_uploads:
        try:
            ChunkedUploadService.cancel_upload(upload.upload_id, upload.user)
            cleaned += 1
        except Exception as e:
            logger.error(f"清理过期上传失败: {upload.upload_id}, error: {e}")
    
    logger.info(f"清理了 {cleaned} 个过期上传")
    return {'cleaned': cleaned}


@shared_task
def cleanup_soft_deleted_files():
    """清理软删除的文件（永久删除）"""
    from .models import FileStorage
    from .services import FileService
    from django.conf import settings
    
    # 删除超过30天的文件
    days = getattr(settings, 'FILE_RETENTION_DAYS', 30)
    cutoff_date = timezone.now() - timezone.timedelta(days=days)
    
    deleted_files = FileStorage.objects.filter(
        is_deleted=True,
        deleted_at__lt=cutoff_date
    )
    
    count = 0
    for file_record in deleted_files:
        try:
            # 从存储中删除
            storage = get_storage_backend(file_record.storage_type)
            if file_record.storage_type == 'oss':
                storage.delete(file_record.oss_key)
            else:
                storage.delete(file_record.storage_path)
            
            # 删除数据库记录
            file_record.delete()
            count += 1
        except Exception as e:
            logger.error(f"清理文件失败: {file_record.id}, error: {e}")
    
    logger.info(f"永久删除了 {count} 个文件")
    return {'deleted': count}


@shared_task
def generate_file_thumbnail(file_id: int):
    """
    生成文件缩略图
    
    Args:
        file_id: 文件ID
    """
    from .models import FileStorage
    from .services import get_storage_backend
    
    try:
        file_record = FileStorage.objects.get(id=file_id)
        
        # 只处理图片
        if not file_record.is_image:
            return {'status': 'skipped', 'reason': 'not_image'}
        
        # 获取原图
        storage = get_storage_backend(file_record.storage_type)
        
        if file_record.storage_type == 'oss':
            content = storage.download(file_record.oss_key)
        else:
            content = storage.download(file_record.storage_path)
        
        # 生成缩略图
        thumbnail_content = create_thumbnail(content, max_size=(200, 200))
        
        # 保存缩略图
        thumbnail_path = file_record.storage_path.replace(
            file_record.file_extension,
            '_thumb.jpg'
        )
        
        storage.upload(thumbnail_content, thumbnail_path)
        
        # 更新文件元数据
        metadata = file_record.metadata or {}
        metadata['thumbnail_path'] = thumbnail_path
        file_record.metadata = metadata
        file_record.save(update_fields=['metadata'])
        
        logger.info(f"缩略图生成完成: {file_id}")
        return {'status': 'success', 'file_id': file_id}
        
    except Exception as e:
        logger.error(f"缩略图生成失败: {file_id}, error: {e}")
        return {'status': 'error', 'reason': str(e)}


@shared_task
def send_file_access_notification(file_id: int, user_id: int, action: str):
    """发送文件访问通知"""
    # TODO: 实现通知功能
    logger.info(f"文件访问通知: file_id={file_id}, user_id={user_id}, action={action}")
    return {'status': 'sent'}


# ==================== 辅助函数 ====================

def extract_text_from_file(content: bytes, file_extension: str) -> str:
    """
    从文件中提取文本内容
    
    Args:
        content: 文件内容
        file_extension: 文件扩展名
        
    Returns:
        提取的文本内容
    """
    ext = file_extension.lower()
    
    # 纯文本文件
    if ext in ['.txt', '.md', '.markdown']:
        try:
            return content.decode('utf-8')
        except UnicodeDecodeError:
            return content.decode('gbk', errors='ignore')
    
    # PDF文件
    elif ext == '.pdf':
        return extract_text_from_pdf(content)
    
    # Word文档
    elif ext in ['.doc', '.docx']:
        return extract_text_from_doc(content, ext)
    
    # Excel文件
    elif ext in ['.xls', '.xlsx']:
        return extract_text_from_excel(content, ext)
    
    # 其他文件暂不支持
    else:
        logger.warning(f"不支持提取文本的文件类型: {ext}")
        return ""


def extract_text_from_pdf(content: bytes) -> str:
    """从PDF提取文本"""
    try:
        import PyPDF2
        from io import BytesIO
        
        pdf_file = BytesIO(content)
        reader = PyPDF2.PdfReader(pdf_file)
        
        text_parts = []
        for page in reader.pages:
            text_parts.append(page.extract_text())
        
        return '\n'.join(text_parts)
    except ImportError:
        logger.warning("PyPDF2未安装，无法提取PDF文本")
        return ""
    except Exception as e:
        logger.error(f"PDF文本提取失败: {e}")
        return ""


def extract_text_from_doc(content: bytes, ext: str) -> str:
    """从Word文档提取文本"""
    try:
        if ext == '.docx':
            from docx import Document
            from io import BytesIO
            
            doc = Document(BytesIO(content))
            return '\n'.join([p.text for p in doc.paragraphs])
        else:
            # .doc格式需要使用python-docx或win32com
            logger.warning(".doc格式暂不支持")
            return ""
    except ImportError:
        logger.warning("python-docx未安装，无法提取Word文本")
        return ""
    except Exception as e:
        logger.error(f"Word文本提取失败: {e}")
        return ""


def extract_text_from_excel(content: bytes, ext: str) -> str:
    """从Excel提取文本"""
    try:
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        
        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"[{sheet_name}]")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = ' '.join([str(cell) if cell is not None else '' for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)
        
        return '\n'.join(text_parts)
    except ImportError:
        logger.warning("openpyxl未安装，无法提取Excel文本")
        return ""
    except Exception as e:
        logger.error(f"Excel文本提取失败: {e}")
        return ""


def call_vector_service(text: str, metadata: dict) -> dict:
    """
    调用向量化服务
    
    Args:
        text: 文本内容
        metadata: 元数据
        
    Returns:
        向量化结果
    """
    # TODO: 根据配置调用实际的向量化服务
    # 这里预留接口，可以接入:
    # - OpenAI Embeddings
    # - Sentence Transformers
    # - 向量数据库服务
    
    vector_id = f"vec_{metadata.get('file_id')}_{timezone.now().strftime('%Y%m%d%H%M%S')}"
    
    logger.info(f"向量化服务预留: vector_id={vector_id}, text_length={len(text)}")
    
    return {
        'vector_id': vector_id,
        'status': 'success'
    }


def create_thumbnail(content: bytes, max_size: tuple = (200, 200)) -> bytes:
    """
    创建缩略图
    
    Args:
        content: 图片内容
        max_size: 最大尺寸
        
    Returns:
        缩略图内容
    """
    try:
        from PIL import Image
        from io import BytesIO
        
        img = Image.open(BytesIO(content))
        
        # 转换为RGB（处理PNG等透明图片）
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # 缩放
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        # 保存为JPEG
        output = BytesIO()
        img.save(output, format='JPEG', quality=85)
        
        return output.getvalue()
        
    except ImportError:
        logger.warning("Pillow未安装，无法生成缩略图")
        return content
    except Exception as e:
        logger.error(f"缩略图生成失败: {e}")
        return content


def get_storage_backend(storage_type: str = None):
    """获取存储后端"""
    from .storage import storage_manager
    return storage_manager.get_storage(storage_type)
